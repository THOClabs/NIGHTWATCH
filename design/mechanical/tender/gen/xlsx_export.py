"""
Vendor XLSX exporter — renders the generated BOM CSVs into a single formatted
Excel workbook a fabricator can quote directly from.

Requires openpyxl (``pip install openpyxl``) — it is NOT a test dependency; the
pure-stdlib CSVs in ``bom/`` are the source of truth and the committed workbook is
a convenience deliverable. Run:  python3 -m design.mechanical.tender.gen.xlsx_export
"""

from __future__ import annotations

import csv
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BOM = pathlib.Path(__file__).resolve().parents[1] / "bom"
OUT = BOM / "NIGHTWATCH_tender_BOM.xlsx"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="A00000")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws, max_w=60):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(12, width + 2))


def _sheet_from_csv(wb, title, csv_path, banner):
    ws = wb.create_sheet(title[:31])
    ws["A1"] = "NIGHTWATCH OBSERVATORY — VENDOR TENDER"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = banner
    ws["A3"] = "ISSUED FOR BID / FOR PE REVIEW — NOT FOR CONSTRUCTION"
    ws["A3"].font = NOTE_FONT
    with csv_path.open() as fh:
        rows = list(csv.reader(fh))
    header_row = 5
    for j, name in enumerate(rows[0], start=1):
        ws.cell(row=header_row, column=j, value=name.replace("_", " ").title())
    _style_header(ws, header_row, len(rows[0]))
    for i, r in enumerate(rows[1:], start=header_row + 1):
        for j, val in enumerate(r, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws)
    return ws


def build() -> pathlib.Path:
    wb = Workbook()
    wb.remove(wb.active)

    idx = wb.create_sheet("Index")
    idx["A1"] = "NIGHTWATCH — Mount & Roll-off 'Turret' Tender BOM"
    idx["A1"].font = TITLE_FONT
    lines = [
        "",
        "This workbook is generated from design/mechanical/tender/bom/*.csv,",
        "which are generated from the proven design/mechanical/calc/params.py.",
        "",
        "Sheets:",
        "  • Master BOM     — every line item, all packages BP-01..BP-05",
        "  • Cut List       — fabricated stock to ORDER & CUT (dual mm / inch)",
        "  • COTS Schedule  — purchased / off-the-shelf items",
        "",
        "Dimensions are dual-unit (mm primary, inch in brackets).",
        "Masses are STOCK mass (what you order), not finished-part net mass.",
        "",
        "STATUS: ISSUED FOR BID / FOR PE REVIEW — NOT FOR CONSTRUCTION.",
        "The pier (BP-03) and roll-off roof (BP-04) require a licensed PE stamp.",
    ]
    for i, ln in enumerate(lines, start=2):
        idx.cell(row=i, column=1, value=ln)
    idx.column_dimensions["A"].width = 80

    _sheet_from_csv(wb, "Master BOM", BOM / "master_bom.csv", "Master Bill of Materials — all packages")
    _sheet_from_csv(wb, "Cut List", BOM / "cut_list.csv", "Cut List / Material Takeoff — fabricated stock")
    _sheet_from_csv(wb, "COTS Schedule", BOM / "cots_schedule.csv", "Purchased / off-the-shelf items")

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print("wrote", build())
